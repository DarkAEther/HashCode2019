import os
import re
import openpyxl
from collections import OrderedDict
from openpyxl.utils.cell import column_index_from_string

TITLE=0
DESCRIPTION=1
ATTACHMENT=2
TEAM_NAME=3
THEME=4 #unused - only 1 theme
STATUS=5
LEADER_NAME=6
LEADER_EMAIL=7


raw_projects =[]
clean_projects = OrderedDict()
raw_projects = openpyxl.load_workbook("hashcode-idea-submissions.xlsx")
sheet = raw_projects.active

final_xlsx = openpyxl.Workbook()
final_sheet = final_xlsx.active
row_count =0
headers =''

join_xlsx = openpyxl.Workbook()
join_projects = join_xlsx.active

teams = openpyxl.load_workbook("hashcode-teams.xlsx")
team_sheet = teams.active

for row in sheet.iter_rows(min_row=1, min_col=1, max_col=8): #iterate over all rows
    #first remove disqualified idea and get the latest idea
    dlist= []
    
    for cell in row:
        dlist.append(cell.value)
    if (row_count== 0): #skip header processing
        row_count+=1
        headers = dlist
        continue
    #print("Processing idea number ",row_count)
    if (dlist[STATUS] != "Disqualified"): #not disqualified
        clean_projects[dlist[TEAM_NAME]] = (tuple(dlist),row_count) #overwrites data, hence only latest idea remains
    row_count+=1


final_sheet.append(headers)
final_clean_projects = sorted(clean_projects.values(),key = lambda x: x[1])
for item in final_clean_projects:
    final_sheet.append(item[0])
#    if (item[0][ATTACHMENT]!=None):
#        print("WILL GET", item[0][ATTACHMENT])

final_xlsx.save("final_ideas.xlsx")

final_headers = []

team_headers = [x.value for x in list(team_sheet.iter_rows())[0]];
final_headers.extend(team_headers)
final_headers.extend(headers[:6])  #excluding repeated headers

join_projects.append(final_headers)   #adding final headers 

c = 0  #no. of ideas

def check_disq(leader_email):
    for row in sheet.iter_rows(min_row =2,min_col = 1,max_col = 8):
        row = [x.value for x in row]
        if(row[-1] == leader_email):
            if(row[5] != 'Disqualified'):
                return 1
            else:
                return 0
  

#flag = 0 #default disqualified

for row in team_sheet.iter_rows(min_row = 2,min_col = 1,max_col = 12):
    row = [x.value for x in row]
    id = []   #idea details list
        
    if(row[3] == 'Team Leader'):
        flag = check_disq(row[1])
        for i in final_sheet.iter_rows(min_row =2,min_col = 1,max_col = 8):
           i = [x.value for x in i]
           if(i[-1] == row[1]): #checking for leader email equality        
                id.extend(i[:6])  
                c+=1 
        row.extend(id)

    if(flag == 1):
            join_projects.append(row)

join_xlsx.save("joined_ideas.xlsx")
print("No. of ideas: ",c)

